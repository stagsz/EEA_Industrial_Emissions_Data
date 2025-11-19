import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_spanish_priority_contacts_excel():
    """Convert Spanish Facilities Priority Contact List CSV to formatted Excel"""

    # Read the CSV file
    csv_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Priority_Contact_List.csv"

    # Read and clean the data
    df = pd.read_csv(csv_path, encoding='utf-8-sig', on_bad_lines='skip')

    # The CSV appears to have duplicate/malformed columns, let's extract the right columns
    # Based on the file structure, we need to parse it carefully

    # Read raw data
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # Parse the header to find the correct columns
    if len(lines) > 1:
        # Clean up the data by taking the visible columns from row 2 onwards
        data_rows = []
        for line in lines[1:]:
            if line.strip():
                # Split by semicolon and parse
                parts = [p.strip() for p in line.split(';')]
                if len(parts) >= 14:  # Ensure we have enough columns
                    data_rows.append(parts)

        # Create dataframe with proper columns
        columns = ['Contact_Roles', 'Approach_Method', 'NOx_Risk', 'Urgent_Compliance',
                   'Region', 'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
                   'Full_Address', 'Primary_Target', 'Company_Type',
                   'Estimated_Contract_Value', 'Decision_Timeline', 'Key_Contacts']

        # Pad columns to match
        max_cols = max(len(row) for row in data_rows)
        for row in data_rows:
            while len(row) < max_cols:
                row.append('')

        df = pd.DataFrame(data_rows[:6])  # Take only the 6 valid data rows

        # Assign proper column names based on the visible structure
        df.columns = ['Contact_Roles', 'Approach_Method', 'NOx_Risk', 'Urgent_Compliance',
                      'Region', 'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
                      'Full_Address', 'Primary_Target', 'Company_Type',
                      'Estimated_Contract_Value', 'Decision_Timeline', 'Key_Contacts'] + \
                     [f'Extra_{i}' for i in range(len(df.columns) - 15)]

        # Keep only the relevant columns
        df = df[['Company_Name', 'City', 'Priority_Level', 'Lead_Score', 'Full_Address',
                 'Primary_Target', 'Company_Type', 'Estimated_Contract_Value',
                 'Decision_Timeline', 'Key_Contacts', 'Approach_Method',
                 'NOx_Risk', 'Urgent_Compliance', 'Region', 'Contact_Roles']]

    # Create output filename
    output_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Priority_Contact_List_Formatted.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Priority Action List - Most important info for immediate action
        action_list = df[[
            'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
            'Estimated_Contract_Value', 'Decision_Timeline',
            'Urgent_Compliance', 'Approach_Method', 'Region'
        ]].copy()
        action_list = action_list.sort_values('Lead_Score', ascending=False)
        action_list.to_excel(writer, sheet_name='Priority Action List', index=False)

        # Sheet 2: Contact Planning - Focus on outreach strategy
        contact_plan = df[[
            'Company_Name', 'City', 'Region', 'Priority_Level',
            'Key_Contacts', 'Contact_Roles', 'Approach_Method',
            'Decision_Timeline', 'Estimated_Contract_Value'
        ]].copy()
        contact_plan = contact_plan.sort_values('Priority_Level')
        contact_plan.to_excel(writer, sheet_name='Contact Planning', index=False)

        # Sheet 3: Complete Information
        df.to_excel(writer, sheet_name='Complete Details', index=False)

        # Sheet 4: By Region - Geographic organization
        regional = df[[
            'Region', 'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
            'Estimated_Contract_Value', 'Decision_Timeline', 'Urgent_Compliance'
        ]].copy()
        regional = regional.sort_values(['Region', 'Lead_Score'], ascending=[True, False])
        regional.to_excel(writer, sheet_name='By Region', index=False)

        # Sheet 5: By Priority Level
        by_priority = df.sort_values(['Priority_Level', 'Lead_Score'], ascending=[True, False])
        by_priority.to_excel(writer, sheet_name='By Priority', index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)

    # Define styles
    critical_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    urgent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column width configurations for each sheet
    column_widths = {
        'Priority Action List': {
            'A': 40, 'B': 20, 'C': 18, 'D': 12, 'E': 22, 'F': 25, 'G': 20, 'H': 30, 'I': 18
        },
        'Contact Planning': {
            'A': 40, 'B': 20, 'C': 18, 'D': 18, 'E': 35, 'F': 45, 'G': 30, 'H': 25, 'I': 22
        },
        'Complete Details': {
            'A': 40, 'B': 20, 'C': 18, 'D': 12, 'E': 45, 'F': 40, 'G': 25,
            'H': 22, 'I': 25, 'J': 35, 'K': 30, 'L': 15, 'M': 20, 'N': 18, 'O': 45
        },
        'By Region': {
            'A': 18, 'B': 40, 'C': 20, 'D': 18, 'E': 12, 'F': 22, 'G': 25, 'H': 20
        },
        'By Priority': {
            'A': 40, 'B': 20, 'C': 18, 'D': 12, 'E': 45, 'F': 40, 'G': 25,
            'H': 22, 'I': 25, 'J': 35, 'K': 30, 'L': 15, 'M': 20, 'N': 18, 'O': 45
        }
    }

    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Set column widths
        if sheet_name in column_widths:
            for col, width in column_widths[sheet_name].items():
                ws.column_dimensions[col].width = width

        # Format data rows with conditional formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Get header name
                header_value = str(ws.cell(1, col_idx).value)

                # Conditional formatting for priority levels
                if 'Priority_Level' in header_value or 'Priority' in header_value:
                    if cell.value == 'CRITICAL':
                        cell.fill = critical_fill
                        cell.font = critical_font
                    elif cell.value == 'HIGH':
                        cell.fill = high_fill
                    elif cell.value == 'MEDIUM':
                        cell.fill = medium_fill

                # Conditional formatting for urgent compliance
                if 'Urgent' in header_value:
                    if str(cell.value).upper() == 'YES':
                        cell.fill = urgent_fill
                        cell.font = Font(bold=True, color="C00000")

                # Number formatting for scores
                if 'Score' in header_value:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0'
                        if cell.value >= 90:
                            cell.font = Font(bold=True, color="C00000")
                        elif cell.value >= 80:
                            cell.font = Font(bold=True, color="FF6B6B")

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Set row height for header
        ws.row_dimensions[1].height = 35

        # Auto-filter on header row
        ws.auto_filter.ref = ws.dimensions

    # Save formatted workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"SUCCESS: Spanish Facilities Priority Contact List formatted successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row-1} rows)")

    # Print priority breakdown
    print(f"\n{'='*80}")
    print(f"PRIORITY CONTACT LIST")
    print(f"{'='*80}")

    for idx, row in df.sort_values('Lead_Score', ascending=False).iterrows():
        print(f"\n{row['Company_Name']} - {row['City']}")
        print(f"  Priority: {row['Priority_Level']} | Score: {row['Lead_Score']}")
        print(f"  Contract Value: {row['Estimated_Contract_Value']}")
        print(f"  Timeline: {row['Decision_Timeline']}")
        print(f"  Urgent Compliance: {row['Urgent_Compliance']}")
        print(f"  Approach: {row['Approach_Method']}")
        print(f"  Region: {row['Region']}")

    # Regional summary
    print(f"\n{'='*80}")
    print(f"REGIONAL SUMMARY")
    print(f"{'='*80}")

    # Convert Lead_Score to numeric for calculations
    df['Lead_Score_Numeric'] = pd.to_numeric(df['Lead_Score'], errors='coerce')

    regional_summary = df.groupby('Region').agg({
        'Company_Name': 'count',
        'Lead_Score_Numeric': 'mean'
    })

    for region, data in regional_summary.iterrows():
        print(f"\n{region}:")
        print(f"  Facilities: {int(data['Company_Name'])}")
        print(f"  Average Score: {data['Lead_Score_Numeric']:.1f}")
        region_facilities = df[df['Region'] == region]['Company_Name'].tolist()
        for facility in region_facilities:
            print(f"    - {facility}")

    print(f"\n{'='*80}\n")

if __name__ == "__main__":
    format_spanish_priority_contacts_excel()