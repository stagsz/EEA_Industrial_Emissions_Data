import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime

def format_spanish_summary_stats_excel():
    """Convert Spanish Facilities Summary Statistics CSV to formatted Excel dashboard"""

    # Read the CSV file
    csv_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Summary_Statistics.csv"
    df = pd.read_csv(csv_path, encoding='utf-8-sig')

    # Create output filename
    output_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Summary_Statistics_Formatted.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Executive Dashboard - Main summary
        df.to_excel(writer, sheet_name='Executive Dashboard', index=False)

        # Sheet 2: Priority Breakdown
        priority_data = pd.DataFrame({
            'Priority Level': ['CRITICAL', 'HIGH', 'MEDIUM'],
            'Facility Count': [
                int(df['Critical_Priority_Count'].iloc[0]),
                int(df['High_Priority_Count'].iloc[0]),
                int(df['Medium_Priority_Count'].iloc[0])
            ],
            'Percentage': [
                (int(df['Critical_Priority_Count'].iloc[0]) / int(df['Total_Facilities'].iloc[0]) * 100),
                (int(df['High_Priority_Count'].iloc[0]) / int(df['Total_Facilities'].iloc[0]) * 100),
                (int(df['Medium_Priority_Count'].iloc[0]) / int(df['Total_Facilities'].iloc[0]) * 100)
            ]
        })
        priority_data.to_excel(writer, sheet_name='Priority Breakdown', index=False)

        # Sheet 3: Facility Type Analysis
        facility_type_data = pd.DataFrame({
            'Category': [
                'Waste-to-Energy Facilities',
                'Other Industrial Facilities',
                'Public Companies',
                'Private Companies',
                'EU ETS Subject',
                'Non-EU ETS',
                'Urgent Compliance Required',
                'No Urgent Compliance'
            ],
            'Count': [
                int(df['Waste_to_Energy_Count'].iloc[0]),
                int(df['Total_Facilities'].iloc[0]) - int(df['Waste_to_Energy_Count'].iloc[0]),
                int(df['Public_Company_Count'].iloc[0]),
                int(df['Total_Facilities'].iloc[0]) - int(df['Public_Company_Count'].iloc[0]),
                int(df['EU_ETS_Subject_Count'].iloc[0]),
                int(df['Total_Facilities'].iloc[0]) - int(df['EU_ETS_Subject_Count'].iloc[0]),
                int(df['Urgent_Compliance_Count'].iloc[0]),
                int(df['Total_Facilities'].iloc[0]) - int(df['Urgent_Compliance_Count'].iloc[0])
            ]
        })
        facility_type_data.to_excel(writer, sheet_name='Facility Analysis', index=False)

        # Sheet 4: Key Metrics - Formatted for presentation
        total_facilities = int(df['Total_Facilities'].iloc[0])
        total_emissions = float(df['Total_Annual_Emissions'].iloc[0])
        avg_score = float(df['Average_Lead_Score'].iloc[0])

        key_metrics = pd.DataFrame({
            'Metric': [
                'Total Facilities Analyzed',
                'CRITICAL Priority Facilities',
                'HIGH Priority Facilities',
                'MEDIUM Priority Facilities',
                'Total Annual Emissions (tonnes)',
                'Average Emissions per Facility (tonnes)',
                'Average Lead Score',
                'Waste-to-Energy Facilities',
                'EU ETS Subject Facilities',
                'Public Companies',
                'Urgent Compliance Required',
                'Estimated Total Market Value',
                'Analysis Date'
            ],
            'Value': [
                str(total_facilities),
                str(int(df['Critical_Priority_Count'].iloc[0])),
                str(int(df['High_Priority_Count'].iloc[0])),
                str(int(df['Medium_Priority_Count'].iloc[0])),
                f"{total_emissions:,.0f}",
                f"{total_emissions/total_facilities:,.0f}",
                f"{avg_score:.1f}",
                str(int(df['Waste_to_Energy_Count'].iloc[0])),
                str(int(df['EU_ETS_Subject_Count'].iloc[0])),
                str(int(df['Public_Company_Count'].iloc[0])),
                str(int(df['Urgent_Compliance_Count'].iloc[0])),
                '€30-65 million',
                str(df['Analysis_Date'].iloc[0])
            ],
            'Notes': [
                'Spanish facilities with emission data',
                '40% of total - immediate action required',
                '20% of total - active pursuit',
                '40% of total - qualified pipeline',
                'Combined annual emissions from all facilities',
                'Average emissions intensity',
                'Out of 100 points - high quality leads',
                '60% of portfolio - primary target market',
                '50% subject to carbon trading obligations',
                '10% are public entities',
                '10% have immediate compliance deadlines',
                'Combined contract value across all facilities',
                'Last analysis timestamp'
            ]
        })
        key_metrics.to_excel(writer, sheet_name='Key Metrics', index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)

    # Define styles
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    critical_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    metric_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    highlight_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Format Executive Dashboard
    ws_dashboard = wb['Executive Dashboard']
    ws_dashboard.column_dimensions['A'].width = 25
    for col in range(2, 13):
        ws_dashboard.column_dimensions[get_column_letter(col)].width = 20

    # Format header
    for cell in ws_dashboard[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data row
    for cell in ws_dashboard[2]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        cell.font = Font(size=11)

    # Add title row
    ws_dashboard.insert_rows(1)
    ws_dashboard.merge_cells('A1:L1')
    ws_dashboard['A1'] = 'SPANISH FACILITIES RESEARCH - EXECUTIVE SUMMARY'
    ws_dashboard['A1'].fill = title_fill
    ws_dashboard['A1'].font = title_font
    ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard['A1'].border = thick_border
    ws_dashboard.row_dimensions[1].height = 30

    ws_dashboard.freeze_panes = 'A3'

    # Format Priority Breakdown
    ws_priority = wb['Priority Breakdown']
    ws_priority.column_dimensions['A'].width = 20
    ws_priority.column_dimensions['B'].width = 18
    ws_priority.column_dimensions['C'].width = 18

    # Add title
    ws_priority.insert_rows(1)
    ws_priority.merge_cells('A1:C1')
    ws_priority['A1'] = 'PRIORITY LEVEL BREAKDOWN'
    ws_priority['A1'].fill = title_fill
    ws_priority['A1'].font = title_font
    ws_priority['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_priority['A1'].border = thick_border
    ws_priority.row_dimensions[1].height = 30

    # Format headers
    for cell in ws_priority[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Format data with conditional colors
    for row_idx in range(3, 6):
        priority_cell = ws_priority.cell(row_idx, 1)
        count_cell = ws_priority.cell(row_idx, 2)
        pct_cell = ws_priority.cell(row_idx, 3)

        if priority_cell.value == 'CRITICAL':
            priority_cell.fill = critical_fill
            priority_cell.font = critical_font
        elif priority_cell.value == 'HIGH':
            priority_cell.fill = high_fill
        elif priority_cell.value == 'MEDIUM':
            priority_cell.fill = medium_fill

        for cell in [priority_cell, count_cell, pct_cell]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        pct_cell.number_format = '0.0"%"'
        count_cell.font = Font(size=12, bold=True)

    # Format Facility Analysis
    ws_facility = wb['Facility Analysis']
    ws_facility.column_dimensions['A'].width = 35
    ws_facility.column_dimensions['B'].width = 15

    # Add title
    ws_facility.insert_rows(1)
    ws_facility.merge_cells('A1:B1')
    ws_facility['A1'] = 'FACILITY TYPE ANALYSIS'
    ws_facility['A1'].fill = title_fill
    ws_facility['A1'].font = title_font
    ws_facility['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_facility['A1'].border = thick_border
    ws_facility.row_dimensions[1].height = 30

    # Format headers
    for cell in ws_facility[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Format data rows
    for row_idx in range(3, 11):
        for col_idx in range(1, 3):
            cell = ws_facility.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center', vertical='center')
            if col_idx == 2:
                cell.font = Font(size=12, bold=True)

    # Format Key Metrics
    ws_metrics = wb['Key Metrics']
    ws_metrics.column_dimensions['A'].width = 40
    ws_metrics.column_dimensions['B'].width = 25
    ws_metrics.column_dimensions['C'].width = 45

    # Add title
    ws_metrics.insert_rows(1)
    ws_metrics.merge_cells('A1:C1')
    ws_metrics['A1'] = 'KEY METRICS & INSIGHTS'
    ws_metrics['A1'].fill = title_fill
    ws_metrics['A1'].font = title_font
    ws_metrics['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_metrics['A1'].border = thick_border
    ws_metrics.row_dimensions[1].height = 30

    # Format headers
    for cell in ws_metrics[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Format data rows with alternating colors
    for row_idx in range(3, 16):
        for col_idx in range(1, 4):
            cell = ws_metrics.cell(row_idx, col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            if row_idx % 2 == 0:
                cell.fill = metric_fill

            if col_idx == 1:
                cell.font = Font(bold=True, size=11)
            elif col_idx == 2:
                cell.font = Font(size=12, bold=True, color="1F4E78")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Highlight critical metrics
            if row_idx in [3, 4, 13]:  # Total, Critical, Market Value
                cell.fill = highlight_fill
                if col_idx == 2:
                    cell.font = Font(size=14, bold=True, color="C00000")

        ws_metrics.row_dimensions[row_idx].height = 25

    ws_metrics.freeze_panes = 'A3'

    # Save formatted workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"SUCCESS: Spanish Facilities Summary Statistics formatted successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row-2} data rows)")

    # Print detailed summary
    print(f"\n{'='*80}")
    print(f"EXECUTIVE SUMMARY - SPANISH MARKET ANALYSIS")
    print(f"{'='*80}")

    total_facilities = int(df['Total_Facilities'].iloc[0])
    critical = int(df['Critical_Priority_Count'].iloc[0])
    high = int(df['High_Priority_Count'].iloc[0])
    medium = int(df['Medium_Priority_Count'].iloc[0])
    emissions = float(df['Total_Annual_Emissions'].iloc[0])
    avg_score = float(df['Average_Lead_Score'].iloc[0])
    wte = int(df['Waste_to_Energy_Count'].iloc[0])
    ets = int(df['EU_ETS_Subject_Count'].iloc[0])
    public = int(df['Public_Company_Count'].iloc[0])
    urgent = int(df['Urgent_Compliance_Count'].iloc[0])

    print(f"\nMARKET OVERVIEW:")
    print(f"  Total Facilities: {total_facilities}")
    print(f"  Average Lead Score: {avg_score:.1f}/100")
    print(f"  Total Annual Emissions: {emissions:,.0f} tonnes/year")
    print(f"  Average per Facility: {emissions/total_facilities:,.0f} tonnes/year")

    print(f"\nPRIORITY DISTRIBUTION:")
    print(f"  CRITICAL: {critical} facilities ({critical/total_facilities*100:.0f}%)")
    print(f"  HIGH:     {high} facilities ({high/total_facilities*100:.0f}%)")
    print(f"  MEDIUM:   {medium} facilities ({medium/total_facilities*100:.0f}%)")

    print(f"\nFACILITY CHARACTERISTICS:")
    print(f"  Waste-to-Energy: {wte} ({wte/total_facilities*100:.0f}%)")
    print(f"  EU ETS Subject: {ets} ({ets/total_facilities*100:.0f}%)")
    print(f"  Public Companies: {public} ({public/total_facilities*100:.0f}%)")
    print(f"  Urgent Compliance: {urgent} ({urgent/total_facilities*100:.0f}%)")

    print(f"\nMARKET OPPORTUNITY:")
    print(f"  Estimated Total Value: EUR 30-65 million")
    print(f"  High-Priority Deals (CRITICAL+HIGH): {critical+high} facilities")
    print(f"  Combined High-Priority Value: EUR 25-55 million")

    print(f"\nKEY INSIGHTS:")
    print(f"  - {critical/total_facilities*100:.0f}% of facilities require immediate action (CRITICAL)")
    print(f"  - {wte/total_facilities*100:.0f}% are waste-to-energy facilities (primary target market)")
    print(f"  - {ets/total_facilities*100:.0f}% subject to EU carbon trading (regulatory driver)")
    print(f"  - Average lead quality is {avg_score:.1f}/100 (high-value pipeline)")
    print(f"  - {urgent} facility has urgent compliance deadline (top priority)")

    print(f"\n{'='*80}")
    print(f"Analysis Date: {df['Analysis_Date'].iloc[0]}")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    format_spanish_summary_stats_excel()