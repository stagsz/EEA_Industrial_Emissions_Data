#!/usr/bin/env python3
"""
Lead Finder - Production Backend Implementation
Automates business lead discovery, qualification, and organization
"""

import json
import re
from dataclasses import dataclass, asdict
from typing import List, Dict, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@dataclass
class Contact:
    """Contact information for a lead"""
    name: str
    title: str
    email: Optional[str] = None
    phone: Optional[str] = None
    linkedin: Optional[str] = None
    email_verified: bool = False

@dataclass
class Lead:
    """Complete lead data structure"""
    id: int
    company: str
    domain: Optional[str]
    industry: str
    location: str
    employees: int
    revenue: float
    contact: Contact
    score: int = 0
    tier: int = 0
    tier_label: str = ""
    recent_news: bool = False
    hiring: bool = False
    growth_indicators: bool = False
    risk_factors: bool = False
    discovery_date: str = ""
    source: str = ""
    notes: str = ""

class LeadScorer:
    """AI-powered lead scoring based on 5-factor model"""
    
    def __init__(self, icp_config: Dict):
        self.icp = icp_config
        
    def calculate_score(self, lead: Lead) -> int:
        """
        Calculate lead score (0-100) based on 5 factors:
        - Company Fit (35% weight, 0-30 points)
        - Budget Potential (25% weight, 0-25 points)
        - Decision Maker Quality (20% weight, 0-20 points)
        - Intent Signals (15% weight, 0-15 points)
        - Risk Factors (-5% weight, -5 points)
        """
        score = 0
        
        # 1. Company Fit (35% - max 30 points)
        company_fit = 0
        
        # Industry match (10 points)
        if lead.industry in self.icp.get('industries', []):
            company_fit += 10
        
        # Company size match (10 points)
        size_min = self.icp.get('company_size', {}).get('min', 0)
        size_max = self.icp.get('company_size', {}).get('max', 999999)
        if size_min <= lead.employees <= size_max:
            company_fit += 10
        
        # Geography match (10 points)
        geographies = self.icp.get('geography', [])
        if any(geo.lower() in lead.location.lower() for geo in geographies):
            company_fit += 10
        
        score += company_fit
        
        # 2. Budget Potential (25% - max 25 points)
        budget_score = 0
        
        # Revenue threshold (15 points)
        revenue_min = self.icp.get('revenue', {}).get('min', 0)
        if lead.revenue >= revenue_min:
            budget_score += 15
            # Bonus for significantly exceeding minimum
            if lead.revenue >= revenue_min * 2:
                budget_score += 5
        
        # Growth indicators (10 points)
        if lead.growth_indicators:
            budget_score += 10
        
        score += min(budget_score, 25)
        
        # 3. Decision Maker Quality (20% - max 20 points)
        dm_score = 0
        
        # Title quality (12 points)
        senior_titles = ['ceo', 'cto', 'cfo', 'vp', 'director', 'head of', 'chief']
        if any(title in lead.contact.title.lower() for title in senior_titles):
            dm_score += 12
        elif 'manager' in lead.contact.title.lower():
            dm_score += 6
        
        # Email availability (8 points)
        if lead.contact.email and '@' in lead.contact.email:
            dm_score += 5
            if lead.contact.email_verified:
                dm_score += 3
        
        score += dm_score
        
        # 4. Intent Signals (15% - max 15 points)
        intent_score = 0
        
        # Recent news/activity (7 points)
        if lead.recent_news:
            intent_score += 7
        
        # Hiring activity (8 points)
        if lead.hiring:
            intent_score += 8
        
        score += intent_score
        
        # 5. Risk Factors (-5%)
        if lead.risk_factors:
            score -= 5
        
        # Ensure score is between 0-100
        return max(0, min(100, score))
    
    def get_tier(self, score: int) -> tuple:
        """Determine lead tier based on score"""
        if score >= 75:
            return (1, "High Priority")
        elif score >= 50:
            return (2, "Qualified")
        else:
            return (3, "Follow-up")

class LeadFinder:
    """Main lead finder orchestrator"""
    
    def __init__(self, icp_config: Dict):
        self.icp = icp_config
        self.scorer = LeadScorer(icp_config)
        self.leads: List[Lead] = []
        
    def process_leads(self, raw_leads: List[Dict]) -> List[Lead]:
        """
        Process and score raw lead data
        
        Args:
            raw_leads: List of dictionaries containing lead data
            
        Returns:
            List of processed and scored Lead objects
        """
        processed_leads = []
        
        for idx, raw_lead in enumerate(raw_leads):
            # Create Contact object
            contact = Contact(
                name=raw_lead.get('contact', {}).get('name', 'Unknown'),
                title=raw_lead.get('contact', {}).get('title', 'Unknown'),
                email=raw_lead.get('contact', {}).get('email'),
                phone=raw_lead.get('contact', {}).get('phone'),
                linkedin=raw_lead.get('contact', {}).get('linkedin'),
                email_verified=raw_lead.get('contact', {}).get('email_verified', False)
            )
            
            # Create Lead object
            lead = Lead(
                id=idx + 1,
                company=raw_lead.get('company', 'Unknown'),
                domain=raw_lead.get('domain'),
                industry=raw_lead.get('industry', 'Unknown'),
                location=raw_lead.get('location', 'Unknown'),
                employees=raw_lead.get('employees', 0),
                revenue=raw_lead.get('revenue', 0),
                contact=contact,
                recent_news=raw_lead.get('recent_news', False),
                hiring=raw_lead.get('hiring', False),
                growth_indicators=raw_lead.get('growth_indicators', False),
                risk_factors=raw_lead.get('risk_factors', False),
                discovery_date=raw_lead.get('discovery_date', datetime.now().strftime('%Y-%m-%d')),
                source=raw_lead.get('source', 'Unknown'),
                notes=raw_lead.get('notes', '')
            )
            
            # Calculate score and tier
            lead.score = self.scorer.calculate_score(lead)
            lead.tier, lead.tier_label = self.scorer.get_tier(lead.score)
            
            processed_leads.append(lead)
            
        self.leads = processed_leads
        return processed_leads
    
    def get_summary_stats(self) -> Dict:
        """Get summary statistics for processed leads"""
        if not self.leads:
            return {
                'total': 0,
                'tier1': 0,
                'tier2': 0,
                'tier3': 0,
                'avg_score': 0
            }
        
        tier1_count = sum(1 for lead in self.leads if lead.tier == 1)
        tier2_count = sum(1 for lead in self.leads if lead.tier == 2)
        tier3_count = sum(1 for lead in self.leads if lead.tier == 3)
        avg_score = sum(lead.score for lead in self.leads) / len(self.leads)
        
        return {
            'total': len(self.leads),
            'tier1': tier1_count,
            'tier2': tier2_count,
            'tier3': tier3_count,
            'avg_score': round(avg_score, 1),
            'tier1_percentage': round((tier1_count / len(self.leads)) * 100, 1),
            'tier2_percentage': round((tier2_count / len(self.leads)) * 100, 1),
            'tier3_percentage': round((tier3_count / len(self.leads)) * 100, 1)
        }
    
    def export_to_excel(self, filename: str = "qualified_leads.xlsx"):
        """
        Export leads to Excel with professional formatting and color-coding
        
        Features:
        - Color-coded by tier (green=Tier 1, yellow=Tier 2, gray=Tier 3)
        - Professional formatting with headers
        - Summary statistics sheet
        - Sortable columns
        """
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: Leads Data
        ws_leads = wb.active
        ws_leads.title = "Qualified Leads"
        
        # Define headers
        headers = [
            "Tier", "Score", "Company", "Domain", "Industry", "Location",
            "Employees", "Revenue ($M)", "Contact Name", "Title",
            "Email", "Phone", "LinkedIn", "Discovery Date", "Source"
        ]
        
        # Write headers with formatting
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_leads.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Define tier colors
        tier_colors = {
            1: PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),  # Green
            2: PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),  # Yellow
            3: PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")   # Gray
        }
        
        # Sort leads by score (highest first)
        sorted_leads = sorted(self.leads, key=lambda x: x.score, reverse=True)
        
        # Write lead data
        for row_num, lead in enumerate(sorted_leads, 2):
            tier_fill = tier_colors[lead.tier]
            
            row_data = [
                f"Tier {lead.tier}",
                lead.score,
                lead.company,
                lead.domain or "N/A",
                lead.industry,
                lead.location,
                lead.employees,
                round(lead.revenue / 1_000_000, 2),
                lead.contact.name,
                lead.contact.title,
                lead.contact.email or "N/A",
                lead.contact.phone or "N/A",
                lead.contact.linkedin or "N/A",
                lead.discovery_date,
                lead.source
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws_leads.cell(row=row_num, column=col_num)
                cell.value = value
                cell.fill = tier_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Bold formatting for score
                if col_num == 2:
                    cell.font = Font(bold=True, size=11)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws_leads.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Sheet 2: Summary Statistics
        ws_summary = wb.create_sheet("Summary")
        stats = self.get_summary_stats()
        
        # Summary header
        ws_summary['A1'] = "Lead Generation Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        ws_summary['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        
        # Summary statistics
        summary_data = [
            ["Metric", "Value"],
            ["Total Leads Discovered", stats['total']],
            ["Tier 1 (High Priority)", f"{stats['tier1']} ({stats['tier1_percentage']}%)"],
            ["Tier 2 (Qualified)", f"{stats['tier2']} ({stats['tier2_percentage']}%)"],
            ["Tier 3 (Follow-up)", f"{stats['tier3']} ({stats['tier3_percentage']}%)"],
            ["Average Lead Score", stats['avg_score']],
            ["Discovery Date", datetime.now().strftime('%Y-%m-%d')],
        ]
        
        for row_num, (metric, value) in enumerate(summary_data, 3):
            ws_summary.cell(row=row_num, column=1, value=metric).font = Font(bold=True)
            ws_summary.cell(row=row_num, column=2, value=value)
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Excel file saved: {filename}")
        print(f"📊 Total leads: {stats['total']}")
        print(f"🥇 Tier 1 leads: {stats['tier1']} ({stats['tier1_percentage']}%)")
        
        return filename

# Example usage
if __name__ == "__main__":
    # Define ICP configuration
    icp_config = {
        'industries': ['Manufacturing', 'Industrial Engineering'],
        'company_size': {'min': 100, 'max': 5000},
        'geography': ['Germany', 'Switzerland', 'Austria'],
        'revenue': {'min': 10_000_000, 'max': 500_000_000}
    }
    
    # Sample raw leads (in production, these would come from web search)
    raw_leads = [
        {
            'company': 'TechForge GmbH',
            'domain': 'techforge.de',
            'industry': 'Manufacturing',
            'location': 'Munich, Germany',
            'employees': 850,
            'revenue': 45_000_000,
            'contact': {
                'name': 'Hans Schmidt',
                'title': 'CEO',
                'email': 'hans.schmidt@techforge.de',
                'linkedin': 'linkedin.com/in/hans-schmidt',
                'email_verified': True
            },
            'recent_news': True,
            'hiring': True,
            'growth_indicators': True,
            'risk_factors': False,
            'source': 'LinkedIn Search'
        },
        {
            'company': 'Alpine Industries AG',
            'domain': 'alpine-industries.ch',
            'industry': 'Industrial Engineering',
            'location': 'Zurich, Switzerland',
            'employees': 1200,
            'revenue': 78_000_000,
            'contact': {
                'name': 'Maria Weber',
                'title': 'VP Operations',
                'email': 'maria.weber@alpine-industries.ch',
                'linkedin': 'linkedin.com/in/maria-weber',
                'email_verified': True
            },
            'recent_news': True,
            'hiring': True,
            'growth_indicators': True,
            'risk_factors': False,
            'source': 'Web Search'
        },
        {
            'company': 'Precision Tools Austria',
            'domain': 'precisiontools.at',
            'industry': 'Manufacturing',
            'location': 'Vienna, Austria',
            'employees': 450,
            'revenue': 23_000_000,
            'contact': {
                'name': 'Klaus Meyer',
                'title': 'Director of Manufacturing',
                'email': 'klaus.meyer@precisiontools.at',
                'email_verified': False
            },
            'recent_news': False,
            'hiring': False,
            'growth_indicators': False,
            'risk_factors': False,
            'source': 'Business Directory'
        }
    ]
    
    # Initialize Lead Finder
    print("🎯 Initializing Lead Finder...")
    finder = LeadFinder(icp_config)
    
    # Process leads
    print("🔍 Processing and scoring leads...")
    processed_leads = finder.process_leads(raw_leads)
    
    # Display results
    print("\n📊 Lead Scoring Results:")
    print("-" * 80)
    for lead in processed_leads:
        print(f"{'🥇' if lead.tier == 1 else '🥈' if lead.tier == 2 else '🥉'} {lead.company}")
        print(f"   Score: {lead.score}/100 | Tier: {lead.tier} ({lead.tier_label})")
        print(f"   Contact: {lead.contact.name} ({lead.contact.title})")
        print(f"   Industry: {lead.industry} | Employees: {lead.employees} | Revenue: ${lead.revenue:,.0f}")
        print()
    
    # Get summary stats
    stats = finder.get_summary_stats()
    print("\n📈 Summary Statistics:")
    print(f"Total Leads: {stats['total']}")
    print(f"Tier 1 (High Priority): {stats['tier1']} ({stats['tier1_percentage']}%)")
    print(f"Tier 2 (Qualified): {stats['tier2']} ({stats['tier2_percentage']}%)")
    print(f"Tier 3 (Follow-up): {stats['tier3']} ({stats['tier3_percentage']}%)")
    print(f"Average Score: {stats['avg_score']}")
    
    # Export to Excel
    print("\n📥 Exporting to Excel...")
    finder.export_to_excel("qualified_leads.xlsx")
    
    print("\n✨ Lead generation complete!")
    print("💡 Next steps:")
    print("   1. Review Tier 1 leads for immediate outreach")
    print("   2. Set up nurture campaigns for Tier 2 leads")
    print("   3. Research additional data for Tier 3 leads")
