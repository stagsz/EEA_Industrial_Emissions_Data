import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_spanish_contacts_excel():
    """Convert Spanish Companies Contact Directory CSV to formatted Excel"""

    # Read the CSV file
    csv_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Companies_Contact_Directory.csv"
    df = pd.read_csv(csv_path, encoding='utf-8-sig')

    # Create output filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Companies_Contact_Directory_Formatted.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main directory to first sheet
        df.to_excel(writer, sheet_name='Contact Directory', index=False)

        # Create company summary sheet
        company_summary = df.groupby('Company_Key').agg({
            'Company_Full_Name': 'first',
            'Company_Type': 'first',
            'Headquarters': 'first',
            'Main_Phone': 'first',
            'Main_Email': 'first',
            'Website': 'first',
            'Procurement_Process': 'first',
            'Decision_Timeline': 'first',
            'Total_Contacts': 'first'
        }).reset_index()

        company_summary.to_excel(writer, sheet_name='Company Overview', index=False)

        # Create contact list by authority level
        contacts_by_authority = df[['Company_Key', 'Company_Full_Name', 'Contact_Title',
                                     'Contact_Department', 'Contact_Phone', 'Contact_Email',
                                     'Contact_Authority', 'Best_Approach']].copy()
        contacts_by_authority = contacts_by_authority.sort_values(['Contact_Authority', 'Company_Key'],
                                                                    ascending=[False, True])
        contacts_by_authority.to_excel(writer, sheet_name='Contacts by Authority', index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    company_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Set column widths and format data cells
        column_widths = {
            'Contact Directory': {
                'A': 15, 'B': 35, 'C': 25, 'D': 40, 'E': 15, 'F': 25,
                'G': 25, 'H': 35, 'I': 30, 'J': 40, 'K': 12, 'L': 12,
                'M': 30, 'N': 25, 'O': 15, 'P': 30, 'Q': 40, 'R': 15, 'S': 35
            },
            'Company Overview': {
                'A': 15, 'B': 35, 'C': 25, 'D': 40, 'E': 15, 'F': 25,
                'G': 25, 'H': 35, 'I': 30, 'J': 12
            },
            'Contacts by Authority': {
                'A': 15, 'B': 35, 'C': 30, 'D': 25, 'E': 15, 'F': 30,
                'G': 15, 'H': 35
            }
        }

        if sheet_name in column_widths:
            for col, width in column_widths[sheet_name].items():
                ws.column_dimensions[col].width = width

        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Highlight company name changes in Contact Directory
                if sheet_name == 'Contact Directory' and cell.column == 1:  # Company_Key column
                    if row_idx == 2 or ws.cell(row_idx, 1).value != ws.cell(row_idx-1, 1).value:
                        for col_cell in row:
                            col_cell.fill = company_fill

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Set row height for header
        ws.row_dimensions[1].height = 30

    # Save formatted workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"SUCCESS: Spanish Companies Contact Directory formatted successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row-1} rows)")

    # Print company summary
    print(f"\n{'='*80}")
    print(f"COMPANY SUMMARY")
    print(f"{'='*80}")

    companies = df.groupby('Company_Key').agg({
        'Company_Full_Name': 'first',
        'Total_Contacts': 'first',
        'Decision_Timeline': 'first'
    })

    for idx, (key, row) in enumerate(companies.iterrows(), 1):
        print(f"\n{idx}. {row['Company_Full_Name']}")
        print(f"   Contacts: {int(row['Total_Contacts'])}")
        print(f"   Decision Timeline: {row['Decision_Timeline']}")

    print(f"\n{'='*80}")
    print(f"Total companies: {len(companies)}")
    print(f"Total contacts: {len(df)}")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    format_spanish_contacts_excel()