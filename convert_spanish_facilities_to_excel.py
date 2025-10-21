import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_spanish_facilities_excel():
    """Convert Spanish Facilities Complete Research Data CSV to formatted Excel"""

    # Read the CSV file (semicolon-delimited)
    csv_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Complete_Research_Data.csv"
    df = pd.read_csv(csv_path, sep=';', encoding='utf-8-sig')

    # Create output filename
    output_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_Complete_Research_Data_Formatted.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Executive Summary - Priority facilities with key metrics
        executive_summary = df[[
            'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
            'Facility_Type', 'Total_Annual_Emissions_Tonnes',
            'NOx_Risk_Level', 'CO2_Risk_Level', 'SO2_Risk_Level',
            'Estimated_Contract_Value', 'Decision_Timeline',
            'Immediate_Services_1'
        ]].copy()
        executive_summary = executive_summary.sort_values('Lead_Score', ascending=False)
        executive_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

        # Sheet 2: Full Research Data
        df.to_excel(writer, sheet_name='Complete Research Data', index=False)

        # Sheet 3: Emissions Analysis
        emissions_data = df[[
            'Company_Name', 'City', 'Priority_Level',
            'Total_Annual_Emissions_Tonnes', 'Number_of_Pollutants',
            'Has_NOx', 'NOx_Emissions_Level', 'NOx_Risk_Level', 'NOx_Concern',
            'Has_CO2', 'CO2_Emissions_Level', 'CO2_Risk_Level', 'CO2_Concern',
            'Has_SO2', 'SO2_Emissions_Level', 'SO2_Risk_Level', 'SO2_Concern',
            'Total_Risk_Count', 'Critical_Risk_Count', 'High_Risk_Count'
        ]].copy()
        emissions_data = emissions_data.sort_values('Total_Annual_Emissions_Tonnes', ascending=False)
        emissions_data.to_excel(writer, sheet_name='Emissions Analysis', index=False)

        # Sheet 4: Compliance & Regulatory
        compliance_data = df[[
            'Company_Name', 'City', 'Priority_Level',
            'Regulatory_Framework', 'BAT_Requirements', 'Monitoring_Requirements',
            'EU_ETS_Subject', 'Urgent_Compliance_Required',
            'NOx_Recommended_Action', 'CO2_Recommended_Action', 'SO2_Recommended_Action',
            'Decision_Timeline', 'Estimated_Contract_Value'
        ]].copy()
        compliance_data = compliance_data.sort_values('Priority_Level')
        compliance_data.to_excel(writer, sheet_name='Compliance & Regulatory', index=False)

        # Sheet 5: Service Opportunities
        service_opportunities = df[[
            'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
            'Estimated_Contract_Value', 'Decision_Timeline',
            'Immediate_Services_1', 'Immediate_Services_2',
            'Requires_SCR_SNCR', 'Requires_Monitoring_Upgrade', 'Requires_BAT_Assessment'
        ]].copy()
        service_opportunities = service_opportunities.sort_values('Lead_Score', ascending=False)
        service_opportunities.to_excel(writer, sheet_name='Service Opportunities', index=False)

        # Sheet 6: Priority Matrix
        priority_summary = df.groupby('Priority_Level').agg({
            'Company_Name': 'count',
            'Lead_Score': 'mean',
            'Total_Annual_Emissions_Tonnes': 'sum',
            'Estimated_Contract_Value': lambda x: list(x),
            'Critical_Risk_Count': 'sum',
            'High_Risk_Count': 'sum'
        }).reset_index()
        priority_summary.columns = ['Priority_Level', 'Facility_Count', 'Avg_Lead_Score',
                                     'Total_Emissions', 'Contract_Values',
                                     'Total_Critical_Risks', 'Total_High_Risks']
        priority_summary.to_excel(writer, sheet_name='Priority Matrix', index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)

    # Define styles
    critical_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column width configurations for each sheet
    column_widths = {
        'Executive Summary': {
            'A': 35, 'B': 20, 'C': 15, 'D': 12, 'E': 30, 'F': 20,
            'G': 15, 'H': 15, 'I': 15, 'J': 20, 'K': 20, 'L': 40
        },
        'Complete Research Data': 'auto',
        'Emissions Analysis': {
            'A': 35, 'B': 20, 'C': 15, 'D': 20, 'E': 15, 'F': 10, 'G': 18, 'H': 15,
            'I': 40, 'J': 10, 'K': 18, 'L': 15, 'M': 40, 'N': 10, 'O': 18, 'P': 15,
            'Q': 40, 'R': 15, 'S': 15, 'T': 15
        },
        'Compliance & Regulatory': {
            'A': 35, 'B': 20, 'C': 15, 'D': 40, 'E': 40, 'F': 40, 'G': 15,
            'H': 20, 'I': 40, 'J': 40, 'K': 40, 'L': 20, 'M': 20
        },
        'Service Opportunities': {
            'A': 35, 'B': 20, 'C': 15, 'D': 12, 'E': 20, 'F': 20, 'G': 40,
            'H': 40, 'I': 18, 'J': 25, 'K': 25
        },
        'Priority Matrix': {
            'A': 18, 'B': 15, 'C': 18, 'D': 18, 'E': 30, 'F': 18, 'G': 18
        }
    }

    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Set column widths
        if sheet_name in column_widths:
            widths = column_widths[sheet_name]
            if widths == 'auto':
                # Auto-size all columns to 15
                for col in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
            else:
                for col, width in widths.items():
                    ws.column_dimensions[col].width = width

        # Format data rows with conditional formatting
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                # Conditional formatting for priority levels
                if 'Priority_Level' in str(ws.cell(1, col_idx).value):
                    if cell.value == 'CRITICAL':
                        cell.fill = critical_fill
                        cell.font = critical_font
                    elif cell.value == 'HIGH':
                        cell.fill = high_fill
                    elif cell.value == 'MEDIUM':
                        cell.fill = medium_fill

                # Conditional formatting for risk levels
                if 'Risk_Level' in str(ws.cell(1, col_idx).value):
                    if cell.value == 'CRITICAL':
                        cell.fill = critical_fill
                        cell.font = critical_font
                    elif cell.value == 'HIGH':
                        cell.fill = high_fill
                    elif cell.value == 'MEDIUM':
                        cell.fill = medium_fill

                # Number formatting
                if 'Emissions' in str(ws.cell(1, col_idx).value) or 'Score' in str(ws.cell(1, col_idx).value):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Set row height for header
        ws.row_dimensions[1].height = 35

    # Save formatted workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"SUCCESS: Spanish Facilities Research Data formatted successfully!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row-1} rows)")

    # Print priority breakdown
    print(f"\n{'='*80}")
    print(f"PRIORITY BREAKDOWN")
    print(f"{'='*80}")

    priority_counts = df['Priority_Level'].value_counts().sort_index()
    for priority, count in priority_counts.items():
        facilities = df[df['Priority_Level'] == priority]['Company_Name'].tolist()
        print(f"\n{priority}: {count} facilities")
        for facility in facilities:
            score = df[df['Company_Name'] == facility]['Lead_Score'].iloc[0]
            print(f"  - {facility} (Score: {score})")

    # Print risk summary
    print(f"\n{'='*80}")
    print(f"RISK SUMMARY")
    print(f"{'='*80}")
    print(f"Total facilities analyzed: {len(df)}")
    print(f"Critical risks identified: {df['Critical_Risk_Count'].sum():.0f}")
    print(f"High risks identified: {df['High_Risk_Count'].sum():.0f}")
    print(f"Total emissions tracked: {df['Total_Annual_Emissions_Tonnes'].sum():,.2f} tonnes/year")

    # Print service opportunities
    print(f"\n{'='*80}")
    print(f"SERVICE OPPORTUNITIES")
    print(f"{'='*80}")

    critical_facilities = df[df['Priority_Level'] == 'CRITICAL'].sort_values('Lead_Score', ascending=False)
    for idx, row in critical_facilities.iterrows():
        print(f"\n{row['Company_Name']} - {row['City']}")
        print(f"  Lead Score: {row['Lead_Score']}")
        print(f"  Contract Value: {row['Estimated_Contract_Value']}")
        print(f"  Timeline: {row['Decision_Timeline']}")
        print(f"  Primary Service: {row['Immediate_Services_1']}")

    print(f"\n{'='*80}\n")

if __name__ == "__main__":
    format_spanish_facilities_excel()
