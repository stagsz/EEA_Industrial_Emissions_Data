import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def format_spanish_complete_excel():
    """Convert Spanish Facilities With Contacts COMPLETE CSV to formatted Excel"""

    # Read the CSV file
    csv_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_With_Contacts_COMPLETE.csv"
    df = pd.read_csv(csv_path, encoding='utf-8-sig')

    # Create output filename
    output_path = r"C:\Users\staff\anthropicFun\EEA_Industrial_Emissions_Data\outputs\Spanish_Facilities_With_Contacts_COMPLETE_Formatted.xlsx"

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Sales Action Dashboard - Most critical info for outreach
        sales_action = df[[
            'Company_Name', 'City', 'Priority_Level', 'Lead_Score',
            'Estimated_Contract_Value', 'Decision_Timeline',
            'Primary_Contact_Title', 'Primary_Contact_Email', 'Primary_Contact_Phone',
            'Primary_Contact_Approach', 'Urgent_Compliance_Required',
            'Immediate_Services_1', 'Region'
        ]].copy()
        sales_action = sales_action.sort_values('Lead_Score', ascending=False)
        sales_action.to_excel(writer, sheet_name='Sales Action Dashboard', index=False)

        # Sheet 2: Complete Contact Directory
        contact_directory = df[[
            'Company_Name', 'Company_Full_Name', 'City', 'Full_Address',
            'Main_Phone', 'Main_Email', 'Company_Website',
            'Primary_Contact_Title', 'Primary_Contact_Email', 'Primary_Contact_Phone',
            'Primary_Contact_Responsibilities', 'Primary_Contact_Authority',
            'Secondary_Contact_Title', 'Secondary_Contact_Email', 'Secondary_Contact_Phone',
            'Plant_Contact_Title', 'Plant_Contact_Email', 'Plant_Contact_Phone',
            'Procurement_Contact_Title', 'Procurement_Contact_Email', 'Procurement_Contact_Phone',
            'All_Key_Contacts', 'Total_Key_Contacts', 'Contact_Research_Status'
        ]].copy()
        contact_directory.to_excel(writer, sheet_name='Complete Contact Directory', index=False)

        # Sheet 3: Technical & Emissions Data
        technical_data = df[[
            'Company_Name', 'City', 'Priority_Level', 'Facility_Type',
            'Total_Annual_Emissions_Tonnes', 'Number_of_Pollutants',
            'NOx_Risk_Level', 'NOx_Emissions_Level', 'NOx_Recommended_Action',
            'CO2_Risk_Level', 'CO2_Emissions_Level', 'CO2_Recommended_Action',
            'SO2_Risk_Level', 'SO2_Emissions_Level', 'SO2_Recommended_Action',
            'Total_Risk_Count', 'Critical_Risk_Count', 'High_Risk_Count',
            'Requires_SCR_SNCR', 'Requires_Monitoring_Upgrade', 'Requires_BAT_Assessment'
        ]].copy()
        technical_data = technical_data.sort_values('Total_Annual_Emissions_Tonnes', ascending=False)
        technical_data.to_excel(writer, sheet_name='Technical & Emissions', index=False)

        # Sheet 4: Compliance & Regulatory
        compliance = df[[
            'Company_Name', 'City', 'Priority_Level', 'Urgent_Compliance_Required',
            'Regulatory_Framework', 'BAT_Requirements', 'Monitoring_Requirements',
            'EU_ETS_Subject', 'Regulatory_Pressure',
            'NOx_Concern', 'CO2_Concern', 'SO2_Concern',
            'Decision_Timeline', 'Estimated_Contract_Value'
        ]].copy()
        compliance = compliance.sort_values('Priority_Level')
        compliance.to_excel(writer, sheet_name='Compliance & Regulatory', index=False)

        # Sheet 5: Business Intelligence
        business_intel = df[[
            'Company_Name', 'City', 'Region', 'Province',
            'Priority_Level', 'Lead_Score', 'Business_Potential_Score',
            'Company_Type', 'Market_Position', 'Funding_Availability',
            'Decision_Process', 'Procurement_Process', 'Decision_Timeline_Details',
            'Cultural_Notes', 'Approach_Method',
            'Estimated_Contract_Value', 'Immediate_Services_1'
        ]].copy()
        business_intel = business_intel.sort_values('Business_Potential_Score', ascending=False)
        business_intel.to_excel(writer, sheet_name='Business Intelligence', index=False)

        # Sheet 6: CRITICAL Priority Facilities (Top Focus)
        critical_facilities = df[df['Priority_Level'] == 'CRITICAL'].copy()
        critical_facilities = critical_facilities[[
            'Company_Name', 'City', 'Lead_Score', 'Urgent_Compliance_Required',
            'Estimated_Contract_Value', 'Decision_Timeline',
            'Primary_Contact_Title', 'Primary_Contact_Email', 'Primary_Contact_Phone',
            'Primary_Contact_Approach',
            'Total_Annual_Emissions_Tonnes', 'NOx_Risk_Level', 'CO2_Risk_Level',
            'Immediate_Services_1', 'All_Key_Contacts'
        ]]
        critical_facilities.to_excel(writer, sheet_name='CRITICAL Priority', index=False)

        # Sheet 7: Contact Research Status
        research_status = df[[
            'Company_Name', 'City', 'Priority_Level',
            'Contact_Research_Status', 'Research_Notes',
            'Total_Key_Contacts', 'All_Key_Contacts',
            'Main_Phone', 'Main_Email', 'Company_Website'
        ]].copy()
        research_status.to_excel(writer, sheet_name='Contact Research Status', index=False)

        # Sheet 8: All Data - Complete dump
        df.to_excel(writer, sheet_name='All Data', index=False)

    # Load workbook for formatting
    wb = load_workbook(output_path)

    # Define styles
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    critical_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    urgent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    urgent_font = Font(bold=True, color="C00000")
    needs_research_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column width configurations
    column_widths = {
        'Sales Action Dashboard': {
            'A': 40, 'B': 20, 'C': 18, 'D': 12, 'E': 22, 'F': 25,
            'G': 30, 'H': 30, 'I': 18, 'J': 35, 'K': 20, 'L': 45, 'M': 18
        },
        'Complete Contact Directory': {
            'A': 40, 'B': 40, 'C': 20, 'D': 45, 'E': 18, 'F': 30, 'G': 25,
            'H': 30, 'I': 30, 'J': 18, 'K': 40, 'L': 18, 'M': 30, 'N': 30,
            'O': 18, 'P': 30, 'Q': 30, 'R': 18, 'S': 30, 'T': 30, 'U': 18,
            'V': 50, 'W': 12, 'X': 25
        },
        'Technical & Emissions': {
            'A': 40, 'B': 20, 'C': 18, 'D': 35, 'E': 20, 'F': 15,
            'G': 15, 'H': 18, 'I': 45, 'J': 15, 'K': 18, 'L': 45,
            'M': 15, 'N': 18, 'O': 45, 'P': 15, 'Q': 15, 'R': 15,
            'S': 18, 'T': 25, 'U': 25
        },
        'Compliance & Regulatory': {
            'A': 40, 'B': 20, 'C': 18, 'D': 25, 'E': 45, 'F': 45,
            'G': 45, 'H': 15, 'I': 18, 'J': 45, 'K': 45, 'L': 45,
            'M': 25, 'N': 22
        },
        'Business Intelligence': {
            'A': 40, 'B': 20, 'C': 18, 'D': 18, 'E': 18, 'F': 12,
            'G': 15, 'H': 25, 'I': 20, 'J': 20, 'K': 30, 'L': 40,
            'M': 30, 'N': 45, 'O': 30, 'P': 22, 'Q': 45
        },
        'CRITICAL Priority': {
            'A': 40, 'B': 20, 'C': 12, 'D': 25, 'E': 22, 'F': 25,
            'G': 30, 'H': 30, 'I': 18, 'J': 40, 'K': 20, 'L': 15,
            'M': 15, 'N': 45, 'O': 60
        },
        'Contact Research Status': {
            'A': 40, 'B': 20, 'C': 18, 'D': 25, 'E': 45, 'F': 12,
            'G': 60, 'H': 18, 'I': 30, 'J': 25
        }
    }

    # Format each sheet (except All Data - too many columns)
    for sheet_name in wb.sheetnames:
        if sheet_name == 'All Data':
            ws = wb[sheet_name]
            # Just format headers for All Data sheet
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            ws.freeze_panes = 'A2'
            continue

        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Set column widths
        if sheet_name in column_widths:
            for col, width in column_widths[sheet_name].items():
                ws.column_dimensions[col].width = width
        else:
            # Auto-size columns to 15 for sheets not in config
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

                header_value = str(ws.cell(1, col_idx).value)

                # Conditional formatting for priority levels
                if 'Priority_Level' in header_value or 'Priority' in header_value:
                    if cell.value == 'CRITICAL':
                        cell.fill = critical_fill
                        cell.font = critical_font
                    elif cell.value == 'HIGH':
                        cell.fill = high_fill
                    elif cell.value == 'MEDIUM':
                        cell.fill = medium_fill

                # Conditional formatting for risk levels
                if 'Risk_Level' in header_value:
                    if cell.value == 'CRITICAL':
                        cell.fill = critical_fill
                        cell.font = critical_font
                    elif cell.value == 'HIGH':
                        cell.fill = high_fill
                    elif cell.value == 'MEDIUM':
                        cell.fill = medium_fill

                # Urgent compliance highlighting
                if 'Urgent' in header_value:
                    if str(cell.value).upper() in ['YES', 'TRUE']:
                        cell.fill = urgent_fill
                        cell.font = urgent_font

                # Research status highlighting
                if 'Research_Status' in header_value:
                    if str(cell.value) == 'REQUIRES_RESEARCH':
                        cell.fill = needs_research_fill
                        cell.font = Font(bold=True, color="FF6B00")

                # Number formatting
                if 'Emissions' in header_value or 'Score' in header_value:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Set row height for header
        ws.row_dimensions[1].height = 35

        # Auto-filter
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    # Save formatted workbook
    wb.save(output_path)

    print(f"\n{'='*80}")
    print(f"SUCCESS: Spanish Facilities With Contacts COMPLETE formatted!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_path}")
    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row-1} data rows)")

    # Print detailed facility breakdown
    print(f"\n{'='*80}")
    print(f"COMPLETE FACILITY & CONTACT INTELLIGENCE")
    print(f"{'='*80}")

    for idx, row in df.sort_values('Lead_Score', ascending=False).iterrows():
        print(f"\n{idx+1}. {row['Company_Name']} - {row['City']}")
        print(f"   Priority: {row['Priority_Level']} | Score: {row['Lead_Score']}")
        print(f"   Contract: {row['Estimated_Contract_Value']} | Timeline: {row['Decision_Timeline']}")
        print(f"   Emissions: {row['Total_Annual_Emissions_Tonnes']:,.0f} tonnes/year")
        print(f"   Urgent: {row['Urgent_Compliance_Required']} | Region: {row['Region']}")

        if pd.notna(row['Primary_Contact_Title']):
            print(f"   PRIMARY CONTACT: {row['Primary_Contact_Title']}")
            print(f"     Email: {row['Primary_Contact_Email']}")
            print(f"     Phone: {row['Primary_Contact_Phone']}")
            print(f"     Approach: {row['Primary_Contact_Approach']}")

        if row['Contact_Research_Status'] == 'REQUIRES_RESEARCH':
            print(f"   ** NEEDS CONTACT RESEARCH **")
            print(f"   Note: {row['Research_Notes']}")

    # Contact completeness summary
    print(f"\n{'='*80}")
    print(f"CONTACT RESEARCH STATUS")
    print(f"{'='*80}")

    complete_contacts = df[df['Contact_Research_Status'].isna()].shape[0]
    needs_research = df[df['Contact_Research_Status'] == 'REQUIRES_RESEARCH'].shape[0]

    print(f"\nComplete contact information: {complete_contacts} facilities")
    print(f"Requires additional research: {needs_research} facilities")

    if needs_research > 0:
        print(f"\nFacilities needing contact research:")
        for idx, row in df[df['Contact_Research_Status'] == 'REQUIRES_RESEARCH'].iterrows():
            print(f"  - {row['Company_Name']} ({row['Priority_Level']} priority)")

    print(f"\n{'='*80}")
    print(f"Total facilities: {len(df)}")
    print(f"Total contacts mapped: {df['Total_Key_Contacts'].sum():.0f}")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    format_spanish_complete_excel()
